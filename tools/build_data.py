#!/usr/bin/env python3
"""
Genera i dati del torneo a partire dal foglio Excel.

  python tools/build_data.py

Legge i 2 fogli 2026 (`Calendario26` + `Gironi26`) da
`Resources/Calend Bocce 2026.xlsx`, ricalcola le classifiche dai risultati
e scrive:
  - data/tournament.js   (window.TOURNAMENT = {...}, usato dal sito)
  - data/tournament.json (stessi dati, per ispezione/debug)

Regola classifiche: Vittorie -> Scontro diretto -> Differenza punti.
I conti (Vinte/Giocate e Differenza) sono ricalcolati dai risultati, non copiati.
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Resources" / "Calend Bocce 2026.xlsx"
YEAR = 2026

# Colori del tema "Sheets" del foglio (per risolvere i fill basati su tema).
THEME = {0: "FFFFFF", 1: "000000", 2: "FFFFFF", 3: "000000",
         4: "4F81BD", 5: "C0504D", 6: "9BBB59", 7: "8064A2",
         8: "4BACC6", 9: "F79646"}

MESI = {1: "gennaio", 2: "febbraio", 3: "marzo", 4: "aprile", 5: "maggio",
        6: "giugno", 7: "luglio", 8: "agosto", 9: "settembre",
        10: "ottobre", 11: "novembre", 12: "dicembre"}

# Fasi KO. L'etichetta nella colonna girone puo' portare l'ordine di tabellone
# (OTT1..OTT8, QUAR1..QUAR4, SEMI1/2); le coppie si formano per ordine crescente
# (QUAR1 = OTT1 vs OTT2, ecc.). Le finali restano FIN 1° / FIN 3°.
KO_PREFIXES = [
    ("OTT", ("ottavi", "Ottavi di finale")),
    ("QUAR", ("quarti", "Quarti di finale")),
    ("SEMI", ("semifinali", "Semifinali")),
    ("FIN3", ("finale3", "Finale 3°/4° posto")),
    ("FIN1", ("finale", "Finale")),
]
PHASE_RANK = {"ottavi": 0, "quarti": 1, "semifinali": 2, "finale3": 3, "finale": 4}


def ko_phase(gir):
    """'OTT3' -> ('ottavi', 'Ottavi di finale', 3); 'FIN 1°' -> ('finale', ..., 0).
    None se non e' una fase KO (es. lettera di girone)."""
    g = str(gir).upper().replace("°", "").replace(" ", "")  # OTT3, QUAR1, FIN1, FIN3
    for prefix, (phase, label) in KO_PREFIXES:
        if g.startswith(prefix):
            num = g[len(prefix):]
            return phase, label, int(num) if num.isdigit() else 0
    return None


def cell_hex(cell):
    """Colore di riempimento di una cella -> hex #rrggbb (None se assente)."""
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return "#FFFFFF"  # nessun riempimento = bianco
    fg = fill.fgColor
    rgb = getattr(fg, "rgb", None)
    if isinstance(rgb, str) and len(rgb) >= 6 and rgb not in ("00000000",):
        return "#" + rgb[-6:].upper()
    theme = getattr(fg, "theme", None)
    if isinstance(theme, int):
        return "#" + THEME.get(theme, "FFFFFF")
    return "#FFFFFF"


def parse_when(raw):
    """'martedì 16/06 20:30' -> dict con label, iso, parti."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = re.search(r"(\d{1,2})/(\d{1,2})(?:\s+(\d{1,2}):(\d{2}))?", s)
    weekday = s.split()[0] if s.split() else ""
    if not m:
        return {"label": s, "weekday": weekday, "iso": None,
                "day": None, "month": None, "time": None}
    day, month = int(m.group(1)), int(m.group(2))
    hh, mm = (int(m.group(3)), int(m.group(4))) if m.group(3) else (None, None)
    iso = None
    if hh is not None:
        iso = datetime(YEAR, month, day, hh, mm).isoformat()
    return {
        "label": s,
        "weekday": weekday,
        "day": day,
        "month": month,
        "monthName": MESI[month],
        "time": f"{hh:02d}:{mm:02d}" if hh is not None else None,
        "iso": iso,
        "dateLabel": f"{day} {MESI[month]}",
    }


def load_groups(wb):
    ws = wb["Gironi26"]
    groups = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if not header or not str(header).strip():
            continue
        letter = str(header).replace("Girone", "").strip().upper()
        color = cell_hex(ws.cell(2, col))  # colore preso dalla prima squadra
        teams = []
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v is not None and str(v).strip():
                teams.append(str(v).strip())
        groups.append({"letter": letter, "name": f"Girone {letter}",
                       "color": color, "teams": teams})
    return groups


def load_matches(wb):
    ws = wb["Calendario26"]
    matches = []
    for r in range(2, ws.max_row + 1):
        when_raw = ws.cell(r, 1).value
        gir = ws.cell(r, 2).value
        s1 = ws.cell(r, 3).value
        s2 = ws.cell(r, 4).value
        p1 = ws.cell(r, 5).value
        p2 = ws.cell(r, 6).value
        gir = str(gir).strip() if gir is not None else ""
        has_teams = bool(s1 and str(s1).strip()) and bool(s2 and str(s2).strip())
        ko = ko_phase(gir)
        if not has_teams and not ko:
            continue  # riga vuota / separatore
        when = parse_when(when_raw)
        played = p1 is not None and p2 is not None and str(p1) != "" and str(p2) != ""
        sc1 = int(round(float(p1))) if played else None
        sc2 = int(round(float(p2))) if played else None
        if ko:
            phase, phase_label, order = ko
        else:
            phase, phase_label, order = "girone", f"Girone {gir}", 0
        winner = None
        if played:
            winner = 1 if sc1 > sc2 else (2 if sc2 > sc1 else 0)
        matches.append({
            "phase": phase,
            "phaseLabel": phase_label,
            "group": gir if phase == "girone" else None,
            "when": when,
            "team1": str(s1).strip() if s1 and str(s1).strip() else None,
            "team2": str(s2).strip() if s2 and str(s2).strip() else None,
            "score1": sc1,
            "score2": sc2,
            "played": played,
            "winner": winner,
            "_order": order,
        })
    # I match KO vanno in ordine di tabellone (OTT1..OTT8, QUAR1..); i gironi
    # restano in ordine di foglio (il calendario ordina comunque per data/ora).
    girone = [m for m in matches if m["phase"] == "girone"]
    ko = [m for m in matches if m["phase"] != "girone"]
    ko.sort(key=lambda m: (PHASE_RANK.get(m["phase"], 9), m["_order"]))
    ordered = girone + ko
    for i, m in enumerate(ordered, 1):
        m["id"] = i
        del m["_order"]
    return ordered


def compute_standings(groups, matches):
    """Classifiche per girone. Regola: Vittorie -> classifica avulsa tra le
    pari-merito (scontri diretti -> differenza generale -> punti fatti) -> alfabetico.
    Per i pari a 2 l'avulsa coincide con lo scontro diretto; per i pari a 3+
    (es. triangolo 1-1-1) si usa la differenza generale, non quella interna."""
    standings = {}
    for g in groups:
        letter = g["letter"]
        rows = {t: {"team": t, "played": 0, "won": 0, "lost": 0,
                    "pf": 0, "pa": 0, "diff": 0} for t in g["teams"]}
        group_matches = [m for m in matches
                         if m["phase"] == "girone" and m["group"] == letter and m["played"]]
        for m in group_matches:
            for team, pf, pa in ((m["team1"], m["score1"], m["score2"]),
                                 (m["team2"], m["score2"], m["score1"])):
                if team not in rows:
                    rows[team] = {"team": team, "played": 0, "won": 0,
                                  "lost": 0, "pf": 0, "pa": 0, "diff": 0}
                rec = rows[team]
                rec["played"] += 1
                rec["pf"] += pf
                rec["pa"] += pa
                rec["diff"] += pf - pa
                if pf > pa:
                    rec["won"] += 1
                else:
                    rec["lost"] += 1

        def internal_wins(team, subset):
            """Vittorie nei soli scontri diretti tra le squadre di 'subset'."""
            w = 0
            for m in group_matches:
                if m["team1"] in subset and m["team2"] in subset:
                    if (m["team1"] == team and m["winner"] == 1) or \
                       (m["team2"] == team and m["winner"] == 2):
                        w += 1
            return w

        import itertools

        # Ordina per vittorie totali; dentro ogni gruppo a pari vittorie applica
        # la classifica avulsa: scontri diretti tra le pari-merito -> differenza
        # GENERALE -> punti fatti -> alfabetico. (Per 2 squadre l'avulsa = scontro
        # diretto; per 3+ in triangolo decide la differenza generale.)
        by_wins = sorted(rows.values(), key=lambda r: -r["won"])
        ordered = []
        for _, grp in itertools.groupby(by_wins, key=lambda r: r["won"]):
            grp = list(grp)
            subset = {r["team"] for r in grp}
            grp.sort(key=lambda r: (-internal_wins(r["team"], subset),
                                    -r["diff"], -r["pf"], r["team"]))
            ordered.extend(grp)
        for i, rec in enumerate(ordered, 1):
            rec["pos"] = i
        standings[letter] = ordered
    return standings


def bump_asset_versions(stamp):
    """Aggiorna ?v=<stamp> su CSS/JS/dati in index.html per invalidare la cache."""
    idx = ROOT / "index.html"
    if not idx.exists():
        return False
    html = idx.read_text(encoding="utf-8")
    pattern = r'((?:href|src)="\.?/?(?:assets/css/styles\.css|assets/js/app\.js|data/tournament\.js))(?:\?v=[^"]*)?(")'
    new_html = re.sub(pattern, lambda m: m.group(1) + "?v=" + stamp + m.group(2), html)
    if new_html != html:
        idx.write_text(new_html, encoding="utf-8")
        return True
    return False


def main():
    if not XLSX.exists():
        sys.exit(f"File non trovato: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    groups = load_groups(wb)
    matches = load_matches(wb)
    standings = compute_standings(groups, matches)

    group_matches = [m for m in matches if m["phase"] == "girone"]
    played = [m for m in group_matches if m["played"]]
    upcoming = [m for m in group_matches if not m["played"] and m["when"] and m["when"]["iso"]]

    data = {
        "meta": {
            "title": "Zibello Arena Bocce",
            "edition": "2026",
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "totalTeams": sum(len(g["teams"]) for g in groups),
            "totalGroups": len(groups),
            "groupMatches": len(group_matches),
            "playedMatches": len(played),
        },
        "groups": groups,
        "matches": matches,
        "standings": standings,
    }

    out_json = ROOT / "data" / "tournament.json"
    out_js = ROOT / "data" / "tournament.js"
    out_json.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    out_js.write_text(
        "// Generato da tools/build_data.py - non modificare a mano.\n"
        "window.TOURNAMENT = " + json.dumps(data, ensure_ascii=False) + ";\n",
        encoding="utf-8")

    # cache-busting: nuova versione asset a ogni build -> niente cache vecchia
    stamp = datetime.now().strftime("%Y%m%d%H%M")
    bumped = bump_asset_versions(stamp)

    # --- riepilogo e verifica ---
    print(f"Squadre: {data['meta']['totalTeams']}  Gironi: {len(groups)}")
    print(f"Partite gironi: {len(group_matches)} ({len(played)} giocate, {len(upcoming)} in programma)")
    ko = [m for m in matches if m["phase"] != "girone"]
    print(f"Partite fase finale: {len(ko)}")
    print("\nClassifiche (ordine calcolato):")
    # Baseline atteso = classifiche calcolate da TUTTI i risultati inseriti finora.
    # Va aggiornato quando nuovi risultati cambiano l'ordine (il check sotto lo segnala).
    expected = {
        "A": ["Gli Allenati", "Gli Zebis", "Wood&Beton", "Beverly INPS"],
        "B": ["I Toghini", "Lord&Chri", "The Doctors", "Apostadores"],
        "C": ["I Cugini", "Chocolate Starfishes", "0Sbat", "La Coppia"],
        "D": ["Mirkae", "Il Boccino della Discordia", "Tironi di Bocce", "BocciAli"],
        "E": ["Bad Boys", "Ricci di Mare", "Gnammestrazzi", "Le Cognate"],
        "F": ["A&M", "Zaccaria", "I Cavalli", "T alla seconda"],
        "G": ["I Diavoli", "Gli Hummarell", "Bocce da Urlo", "I Masalén"],
        "H": ["Team Rocket", "Ghirarda", "Atletico Cavalclown 2.0", "Le Sbocciate"],
    }
    ok_all = True
    for letter in sorted(standings):
        order = [r["team"] for r in standings[letter]]
        exp = expected.get(letter)
        ok = (order == exp) if exp else None
        flag = "OK " if ok else ("XXX" if ok is False else "  ?")
        if ok is False:
            ok_all = False
        print(f"  Girone {letter}: {flag} " +
              " | ".join(f"{r['pos']}.{r['team']}({r['won']}/{r['played']} {r['diff']:+d})"
                         for r in standings[letter]))
    print("\nVERIFICA ordine vs baseline atteso:",
          "TUTTO COERENTE" if ok_all else "DISCREPANZE PRESENTI (aggiornare il baseline)")
    print(f"\nScritto: {out_js.relative_to(ROOT)} , {out_json.relative_to(ROOT)}")
    print("Cache-busting index.html:", ("?v=" + stamp) if bumped else "nessuna modifica")


if __name__ == "__main__":
    main()
